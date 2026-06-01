import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUNAS_NECESSARIAS = {"Subtipo", "Data Inclusão", "Numero", "Valor Faturas"}
SUBTIPOS_P1 = [
    "P1 SUSPENSÃO - GRUPO A",
    "P1 SUSPENSÃO - POSTE",
    "P1 VISTORIA - RETIRADA DE RAMAL",
]


def detectar_aba(xls: pd.ExcelFile) -> str | None:
    """Retorna o nome da primeira aba que contenha as colunas necessárias."""
    for sheet in xls.sheet_names:
        try:
            df = xls.parse(sheet, nrows=5)
            if COLUNAS_NECESSARIAS.issubset(df.columns):
                return sheet
        except Exception:
            continue
    return None


def gerar_excel(tabela: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Acomp. P1"

    AZUL_HEADER = "2E5F9E"
    AZUL_TITULO = "1F4E79"
    BRANCO      = "FFFFFF"
    CINZA_LINHA = "DEEAF1"

    thin = Side(style="thin", color="AAAAAA")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Acompanhamento de Suspensão/Vistoria P1"
    c.font = Font(name="Arial", bold=True, color=BRANCO, size=13)
    c.fill = PatternFill("solid", fgColor=AZUL_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Cabeçalhos
    headers = ["DATA", "P1 SUSPENSÃO - GRUPO A", "P1 SUSPENSÃO - POSTE",
               "P1 VISTORIA - RETIRADA DE RAMAL", "Total Geral", "Total Dívida"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = Font(name="Arial", bold=True, color=BRANCO, size=10)
        c.fill = PatternFill("solid", fgColor=AZUL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda
    ws.row_dimensions[2].height = 30

    meses = {"Jan":"jan","Feb":"fev","Mar":"mar","Apr":"abr","May":"mai","Jun":"jun",
             "Jul":"jul","Aug":"ago","Sep":"set","Oct":"out","Nov":"nov","Dec":"dez"}

    for i, row in enumerate(tabela.itertuples(index=False), start=3):
        fill_cor = PatternFill("solid", fgColor=CINZA_LINHA if i % 2 == 0 else BRANCO)

        data_str = row[0].strftime("%d/%b")
        for en, pt in meses.items():
            data_str = data_str.replace(en, pt)

        c = ws.cell(row=i, column=1, value=data_str)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")
        c.fill = fill_cor
        c.border = borda

        for col_idx, val in enumerate([row[1], row[2], row[3], row[4]], start=2):
            c = ws.cell(row=i, column=col_idx, value=int(val))
            c.font = Font(name="Arial", size=10, bold=(col_idx == 5))
            c.alignment = Alignment(horizontal="center")
            c.fill = fill_cor
            c.border = borda

        c = ws.cell(row=i, column=6, value=row[5])
        c.number_format = 'R$ #,##0.00'
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="right")
        c.fill = fill_cor
        c.border = borda

    # Linha Total Geral
    total_row = ws.max_row + 1
    data_start, data_end = 3, total_row - 1
    ws.cell(row=total_row, column=1, value="Total Geral").font = Font(name="Arial", bold=True, color=BRANCO, size=10)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=AZUL_HEADER)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=1).border = borda

    for col_idx, col_letter in enumerate(["B","C","D","E","F"], start=2):
        c = ws.cell(row=total_row, column=col_idx)
        c.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        c.font = Font(name="Arial", bold=True, color=BRANCO, size=10)
        c.fill = PatternFill("solid", fgColor=AZUL_HEADER)
        c.alignment = Alignment(horizontal="center" if col_idx < 6 else "right")
        c.border = borda
        if col_idx == 6:
            c.number_format = 'R$ #,##0.00'

    for i, w in enumerate([12, 26, 22, 34, 14, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── UI ──────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Acompanhamento Suspensão/Vistoria P1", layout="wide")
st.title("📋 Acompanhamento de Suspensão/Vistoria P1")

uploaded_file = st.file_uploader(
    "Suba a base (.xlsx)",
    type=["xlsx"],
    help="Apenas arquivos no formato .xlsx são aceitos.",
)

if uploaded_file is not None:
    # Validar extensão explicitamente
    if not uploaded_file.name.lower().endswith(".xlsx"):
        st.error("⚠️ Formato inválido! Por favor, suba apenas arquivos **.xlsx**.")
        st.stop()

    try:
        xls = pd.ExcelFile(uploaded_file)
        aba = detectar_aba(xls)

        if aba is None:
            abas = ", ".join(xls.sheet_names)
            st.error(
                f"❌ Nenhuma aba válida encontrada no arquivo.\n\n"
                f"Abas encontradas: **{abas}**\n\n"
                f"O arquivo precisa ter as colunas: {', '.join(sorted(COLUNAS_NECESSARIAS))}"
            )
            st.stop()

        st.success(f"✅ Base carregada com sucesso! Aba utilizada: **{aba}**")

        df = xls.parse(aba)

        # Filtrar subtipos P1
        df_p1 = df[df["Subtipo"].isin(SUBTIPOS_P1)].copy()
        df_p1["Data Inclusão"] = pd.to_datetime(df_p1["Data Inclusão"]).dt.date

        # Pivot: contagem por dia e subtipo
        pivot_qtd = (
            df_p1.groupby(["Data Inclusão", "Subtipo"])["Numero"]
            .count()
            .unstack(fill_value=0)
            .reindex(columns=SUBTIPOS_P1, fill_value=0)
        )
        pivot_qtd["Total Geral"] = pivot_qtd.sum(axis=1)

        # Dívida por dia
        divida_dia = (
            df_p1.groupby("Data Inclusão")["Valor Faturas"]
            .sum()
            .rename("Total Dívida")
        )

        tabela = pivot_qtd.join(divida_dia).reset_index().rename(columns={"Data Inclusão": "DATA"}).sort_values("DATA")

        # ── Métricas ──
        st.subheader("Resumo do Período")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("P1 Suspensão - Grupo A",   int(pivot_qtd["P1 SUSPENSÃO - GRUPO A"].sum()))
        col2.metric("P1 Suspensão - Poste",      int(pivot_qtd["P1 SUSPENSÃO - POSTE"].sum()))
        col3.metric("P1 Vistoria - Ret. Ramal",  int(pivot_qtd["P1 VISTORIA - RETIRADA DE RAMAL"].sum()))
        total_fmt = f"R$ {divida_dia.sum():,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        col4.metric("Total Dívida", total_fmt)

        st.divider()

        # ── Tabela + botão download ──
        col_titulo, col_btn = st.columns([4, 1])
        col_titulo.subheader("Detalhamento por Dia")
        excel_bytes = gerar_excel(tabela)
        col_btn.download_button(
            label="⬇️ Baixar Excel",
            data=excel_bytes,
            file_name="acompanhamento_suspensao_p1.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        # Tabela formatada para exibição
        tabela_exibir = tabela.copy()
        meses = {"Jan":"jan","Feb":"fev","Mar":"mar","Apr":"abr","May":"mai","Jun":"jun",
                 "Jul":"jul","Aug":"ago","Sep":"set","Oct":"out","Nov":"nov","Dec":"dez"}
        tabela_exibir["DATA"] = tabela_exibir["DATA"].apply(
            lambda x: "".join(
                meses.get(x.strftime("%b"), x.strftime("%b")) if i == 1 else p
                for i, p in enumerate(x.strftime("%d/%b").split("/"))
            ) if hasattr(x, "strftime") else str(x)
        )
        tabela_exibir["DATA"] = tabela_exibir["DATA"].apply(
            lambda x: x[:3] + "/" + "".join(meses.get(x[3:], x[3:])) if hasattr(x, "split") else x
        )
        tabela_exibir["DATA"] = tabela["DATA"].apply(
            lambda x: x.strftime("%d/") + meses.get(x.strftime("%b"), x.strftime("%b"))
            if hasattr(x, "strftime") else str(x)
        )
        tabela_exibir["Total Dívida"] = tabela["Total Dívida"].apply(
            lambda x: f"R$ {float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        )

        # Linha Total Geral
        totais = tabela.drop(columns="DATA").sum()
        totais_fmt = totais.copy()
        totais_fmt["Total Dívida"] = f"R$ {totais['Total Dívida']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        totais_row = pd.DataFrame([["Total Geral"] + totais_fmt.tolist()], columns=tabela_exibir.columns)
        tabela_exibir = pd.concat([tabela_exibir, totais_row], ignore_index=True)

        st.dataframe(
            tabela_exibir,
            use_container_width=True,
            hide_index=True,
            column_config={
                "DATA": st.column_config.TextColumn("DATA"),
                "P1 SUSPENSÃO - GRUPO A": st.column_config.NumberColumn("P1 SUSPENSÃO - GRUPO A"),
                "P1 SUSPENSÃO - POSTE": st.column_config.NumberColumn("P1 SUSPENSÃO - POSTE"),
                "P1 VISTORIA - RETIRADA DE RAMAL": st.column_config.NumberColumn("P1 VISTORIA - RETIRADA DE RAMAL"),
                "Total Geral": st.column_config.NumberColumn("Total Geral"),
                "Total Dívida": st.column_config.TextColumn("Total Dívida"),
            },
        )

    except Exception as e:
        st.error(f"Erro ao processar o arquivo: {e}")

else:
    st.info("👆 Suba um arquivo .xlsx para começar.")
